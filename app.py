import pdfplumber
import re
import pandas as pd
from fuzzywuzzy import fuzz
import nltk
from nltk.tokenize import word_tokenize
import sys
import time
import logging
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QFileDialog, QTableWidget, QTableWidgetItem, QHeaderView, QProgressBar,
                             QLineEdit, QMessageBox, QFrame)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QPalette, QColor

# Logging Setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler("trademark_detector.log"), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

# Download NLTK data
try:
    nltk.download('punkt', quiet=True)
except Exception as e:
    logger.error(f"Failed to download NLTK data: {e}")
    sys.exit(1)

# Styles Class
class Styles:
    DARK_THEME = {
        "window": "#1e1e1e",
        "text": "#cccccc",
        "button": "#0078d4",
        "button_hover": "#005bb5",
        "process": "#00c4b4",
        "process_hover": "#009688",
        "export": "#ff6f61",
        "export_hover": "#e65a50",
        "highlight": "FFFF99",
        "background": "#2a2a2a",
        "header": "#333"
    }
    FONT = QFont("Arial", 10)
    TITLE_FONT = QFont("Arial", 16, QFont.Bold)
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Utility Functions
def validate_file(file_path, expected_extension):
    """Validate file path and extension."""
    try:
        file_path = Path(file_path).resolve()
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        if not file_path.suffix.lower() == expected_extension.lower():
            raise ValueError(f"Expected a {expected_extension} file, got {file_path.suffix}")
        return file_path
    except Exception as e:
        logger.error(f"File validation failed: {e}")
        raise

# Processing Class
class TrademarkProcessor(QThread):
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(pd.DataFrame)
    error = pyqtSignal(str)

    def __init__(self, tmj_path, portfolio_path):
        super().__init__()
        self.tmj_path = tmj_path
        self.portfolio_path = portfolio_path
        self._is_running = True

    def stop(self):
        self._is_running = False

    def run(self):
        try:
            logger.info("Starting trademark processing")
            self.progress.emit(30, "Extracting trademarks from TMJ PDF...")
            tmj_data = self.extract_trademarks(self.tmj_path)
            if not self._is_running or tmj_data.empty:
                self.error.emit("No trademarks extracted or processing canceled")
                return

            self.progress.emit(60, "Loading portfolio data...")
            portfolio_data = self.load_portfolio(self.portfolio_path)
            if not self._is_running or portfolio_data.empty:
                self.error.emit("Portfolio data is empty or processing canceled")
                return

            self.progress.emit(90, "Comparing trademarks...")
            report_data = self.compare_trademarks(tmj_data, portfolio_data)
            if not self._is_running:
                self.error.emit("Processing canceled")
                return

            self.progress.emit(100, "Processing complete")
            self.finished.emit(report_data)

        except Exception as e:
            logger.error(f"Processing failed: {e}")
            self.error.emit(f"An error occurred: {str(e)}")

    def extract_trademarks(self, pdf_path):
        """Extract all trademarks from the entire TMJ PDF."""
        start_time = time.time()
        trademarks = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
                logger.info(f"Processing PDF with {total_pages} pages")

                for page_num, page in enumerate(pdf.pages, start=1):
                    if not self._is_running:
                        break
                    text = page.extract_text()
                    if not text:
                        logger.warning(f"Page {page_num} is empty, skipping")
                        continue
                    if "ZEUSIP" in text.upper():
                        logger.info(f"Skipping page {page_num} due to 'ZEUSIP' reference")
                        continue

                    pattern = r"(Class \d+)\n(.+?)\n(\d{7,8}) (\d{2}/\d{2}/\d{4})\n(.+?)(?:\nProposed to be Used|Used Since.+?)?\n(.+?)(?=\nClass|\Z)"
                    matches = re.findall(pattern, text, re.DOTALL)

                    for idx, match in enumerate(matches, start=1):
                        trademark_name = match[1].strip()
                        status_match = re.search(r"(Proposed to be Used|Used Since.+?)(?=\n|$)", match[0] + match[5])
                        status = status_match.group(1) if status_match else "Unknown"
                        class_goods = f"{match[0]}: {match[5].strip()}"

                        trademarks.append({
                            "Applicant": match[4].strip(),
                            "Trademark": trademark_name,
                            "Application number": match[2],
                            "Status": status,
                            "Class Goods And Use": class_goods
                        })

            if not trademarks:
                logger.warning("No trademarks extracted from TMJ PDF")
            logger.info(f"Extraction completed in {time.time() - start_time:.2f} seconds")
            return pd.DataFrame(trademarks)

        except Exception as e:
            logger.error(f"Failed to extract trademarks: {e}")
            raise

    def load_portfolio(self, portfolio_path):
        """Load and validate portfolio Excel file."""
        try:
            portfolio_data = pd.read_excel(portfolio_path)
            expected_headers = ["File number", "Lawyer", "Applicant", "Trademark", "Application number", "Status", "Class Goods And Use"]
            if not all(header in portfolio_data.columns for header in expected_headers):
                raise ValueError(f"Portfolio Excel must contain all required headers: {expected_headers}")
            return portfolio_data
        except Exception as e:
            logger.error(f"Failed to load portfolio data: {e}")
            raise

    def compare_trademarks(self, tmj_data, portfolio_data, name_threshold=80, desc_threshold=50):
        """Compare trademarks between TMJ and portfolio."""
        start_time = time.time()
        report_data = portfolio_data.copy()
        report_data["Conflict Marks"] = ""
        report_data["Other Details"] = ""

        conflicts = {idx: [] for idx in range(len(portfolio_data))}

        if tmj_data.empty:
            logger.warning("No TMJ trademarks to compare")
            return report_data

        for tmj_idx, tmj_row in tmj_data.iterrows():
            for portfolio_idx, portfolio_row in portfolio_data.iterrows():
                if not self._is_running:
                    break
                try:
                    name_similarity = fuzz.token_sort_ratio(tmj_row["Trademark"], portfolio_row["Trademark"])
                    tmj_class = tmj_row["Class Goods And Use"].split(":")[0].strip()
                    portfolio_class = portfolio_row["Class Goods And Use"].split(":")[0].strip()
                    class_match = tmj_class == portfolio_class

                    tmj_desc = tmj_row["Class Goods And Use"].split(":", 1)[1].strip().lower()
                    portfolio_desc = portfolio_row["Class Goods And Use"].split(":", 1)[1].strip().lower()
                    tmj_tokens = set(word_tokenize(tmj_desc))
                    portfolio_tokens = set(word_tokenize(portfolio_desc))
                    common_tokens = len(tmj_tokens.intersection(portfolio_tokens))
                    desc_similarity = (common_tokens / max(len(tmj_tokens), len(portfolio_tokens))) * 100 if tmj_tokens or portfolio_tokens else 0

                    if name_similarity > name_threshold or (class_match and desc_similarity > desc_threshold):
                        conflict_detail = (f"{tmj_row['Trademark']} (App No: {tmj_row['Application number']}, "
                                         f"Applicant: {tmj_row['Applicant']}, Status: {tmj_row['Status']}, "
                                         f"Class Goods And Use: {tmj_row['Class Goods And Use']}, "
                                         f"Name Similarity: {name_similarity}%, Description Similarity: {desc_similarity}%)")
                        conflicts[portfolio_idx].append((tmj_row["Trademark"], conflict_detail))

                except Exception as e:
                    logger.error(f"Error comparing TMJ entry {tmj_idx} with portfolio entry {portfolio_idx}: {e}")
                    continue

            if not self._is_running:
                break

        for idx in conflicts:
            if conflicts[idx]:
                conflict_marks = ", ".join([mark for mark, _ in conflicts[idx]])
                other_details = "; ".join([detail for _, detail in conflicts[idx]])
                report_data.at[idx, "Conflict Marks"] = conflict_marks
                report_data.at[idx, "Other Details"] = other_details

        logger.info(f"Comparison completed in {time.time() - start_time:.2f} seconds")
        return report_data

# UI Class
class TrademarkApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Trademark Similarity Detector")
        self.setGeometry(100, 100, 1200, 700)

        # Apply dark theme
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(Styles.DARK_THEME["window"]))
        palette.setColor(QPalette.WindowText, Qt.white)
        palette.setColor(QPalette.Base, QColor(Styles.DARK_THEME["background"]))
        palette.setColor(QPalette.AlternateBase, QColor(Styles.DARK_THEME["header"]))
        palette.setColor(QPalette.Text, Qt.white)
        palette.setColor(QPalette.Button, QColor(60, 60, 60))
        palette.setColor(QPalette.ButtonText, Qt.white)
        palette.setColor(QPalette.Highlight, QColor(0, 120, 215))
        palette.setColor(QPalette.HighlightedText, Qt.white)
        self.setPalette(palette)

        # Main widget and layout
        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)
        self.layout = QHBoxLayout(self.main_widget)

        # Sidebar
        self.sidebar = QFrame()
        self.sidebar.setStyleSheet(f"background-color: {Styles.DARK_THEME['window']}; border-right: 1px solid #333;")
        self.sidebar_layout = QVBoxLayout(self.sidebar)
        self.sidebar_layout.setAlignment(Qt.AlignTop)

        # Title
        self.title_label = QLabel("Trademark Similarity Detector")
        self.title_label.setFont(Styles.TITLE_FONT)
        self.title_label.setStyleSheet("color: #00b4d8; padding: 10px;")
        self.sidebar_layout.addWidget(self.title_label)

        # TMJ Section
        self.tmj_label = QLabel("TMJ PDF: Not selected")
        self.tmj_label.setStyleSheet(f"color: {Styles.DARK_THEME['text']}; padding: 5px;")
        self.sidebar_layout.addWidget(self.tmj_label)

        self.upload_tmj_btn = QPushButton("Upload TMJ PDF")
        self.upload_tmj_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Styles.DARK_THEME['button']}; color: white; padding: 10px; border-radius: 5px;
            }}
            QPushButton:hover {{ background-color: {Styles.DARK_THEME['button_hover']}; }}
        """)
        self.upload_tmj_btn.clicked.connect(self.upload_tmj)
        self.sidebar_layout.addWidget(self.upload_tmj_btn)

        # Portfolio Section
        self.portfolio_label = QLabel("Portfolio Excel: Not selected")
        self.portfolio_label.setStyleSheet(f"color: {Styles.DARK_THEME['text']}; padding: 5px;")
        self.sidebar_layout.addWidget(self.portfolio_label)

        self.upload_portfolio_btn = QPushButton("Upload Portfolio Excel")
        self.upload_portfolio_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Styles.DARK_THEME['button']}; color: white; padding: 10px; border-radius: 5px;
            }}
            QPushButton:hover {{ background-color: {Styles.DARK_THEME['button_hover']}; }}
        """)
        self.upload_portfolio_btn.clicked.connect(self.upload_portfolio)
        self.sidebar_layout.addWidget(self.upload_portfolio_btn)

        # Process Section
        self.process_btn = QPushButton("Process Files")
        self.process_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Styles.DARK_THEME['process']}; color: white; padding: 10px; border-radius: 5px;
            }}
            QPushButton:hover {{ background-color: {Styles.DARK_THEME['process_hover']}; }}
        """)
        self.process_btn.clicked.connect(self.process_files)
        self.process_btn.setEnabled(False)
        self.sidebar_layout.addWidget(self.process_btn)

        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545; color: white; padding: 10px; border-radius: 5px;
            }
            QPushButton:hover { background-color: #c82333; }
        """)
        self.cancel_btn.clicked.connect(self.cancel_processing)
        self.cancel_btn.setEnabled(False)
        self.sidebar_layout.addWidget(self.cancel_btn)

        self.clear_btn = QPushButton("Clear")
        self.clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d; color: white; padding: 10px; border-radius: 5px;
            }
            QPushButton:hover { background-color: #5a6268; }
        """)
        self.clear_btn.clicked.connect(self.clear_ui)
        self.sidebar_layout.addWidget(self.clear_btn)

        # Progress
        self.progress_label = QLabel("Ready")
        self.progress_label.setStyleSheet(f"color: {Styles.DARK_THEME['text']}; padding: 5px;")
        self.sidebar_layout.addWidget(self.progress_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #555; border-radius: 5px; background-color: #333; text-align: center;
            }
            QProgressBar::chunk {
                background-color: #00c4b4;
            }
        """)
        self.progress_bar.setValue(0)
        self.sidebar_layout.addWidget(self.progress_bar)

        self.layout.addWidget(self.sidebar, 1)

        # Main Content
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)

        self.filter_label = QLabel("Filter Results:")
        self.filter_label.setStyleSheet(f"color: {Styles.DARK_THEME['text']};")
        self.content_layout.addWidget(self.filter_label)

        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Enter keyword to filter...")
        self.filter_input.setStyleSheet("background-color: #333; color: white; padding: 5px; border: 1px solid #555; border-radius: 5px;")
        self.filter_input.textChanged.connect(self.filter_table)
        self.content_layout.addWidget(self.filter_input)

        self.table = QTableWidget()
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {Styles.DARK_THEME['background']}; color: white; border: 1px solid #555;
            }}
            QTableWidget::item {{ padding: 5px; }}
            QHeaderView::section {{ background-color: {Styles.DARK_THEME['header']}; color: white; padding: 5px; border: 1px solid #555; }}
        """)
        self.table.setSortingEnabled(True)
        self.content_layout.addWidget(self.table)

        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Styles.DARK_THEME['export']}; color: white; padding: 10px; border-radius: 5px;
            }}
            QPushButton:hover {{ background-color: {Styles.DARK_THEME['export_hover']}; }}
        """)
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)
        self.content_layout.addWidget(self.export_btn)

        self.layout.addWidget(self.content_widget, 4)

        # State
        self.tmj_path = None
        self.portfolio_path = None
        self.results = None
        self.processor = None

    def upload_tmj(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "Select TMJ PDF", "", "PDF Files (*.pdf)")
            if file_path:
                file_path = validate_file(file_path, '.pdf')
                self.tmj_path = str(file_path)
                self.tmj_label.setText(f"TMJ PDF: {file_path.name}")
                self.check_process_button()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def upload_portfolio(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "Select Portfolio Excel", "", "Excel Files (*.xlsx)")
            if file_path:
                file_path = validate_file(file_path, '.xlsx')
                self.portfolio_path = str(file_path)
                self.portfolio_label.setText(f"Portfolio Excel: {file_path.name}")
                self.check_process_button()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def check_process_button(self):
        self.process_btn.setEnabled(bool(self.tmj_path and self.portfolio_path))

    def process_files(self):
        if not self.tmj_path or not self.portfolio_path:
            QMessageBox.critical(self, "Error", "Please upload both TMJ PDF and Portfolio Excel.")
            return

        self.set_ui_processing_state(True)
        self.progress_label.setText("Starting...")
        self.progress_bar.setValue(0)

        self.processor = TrademarkProcessor(self.tmj_path, self.portfolio_path)
        self.processor.progress.connect(self.update_progress)
        self.processor.finished.connect(self.display_results)
        self.processor.error.connect(self.show_error)
        self.processor.start()

    def cancel_processing(self):
        if self.processor and self.processor.isRunning():
            self.processor.stop()
            self.progress_label.setText("Canceling...")
            self.processor.wait()
            self.set_ui_processing_state(False)

    def set_ui_processing_state(self, processing):
        self.process_btn.setEnabled(not processing)
        self.cancel_btn.setEnabled(processing)
        self.upload_tmj_btn.setEnabled(not processing)
        self.upload_portfolio_btn.setEnabled(not processing)
        self.export_btn.setEnabled(not processing and self.results is not None)
        self.clear_btn.setEnabled(not processing)
        QApplication.setOverrideCursor(Qt.WaitCursor if processing else Qt.ArrowCursor)

    def clear_ui(self):
        self.tmj_path = None
        self.portfolio_path = None
        self.results = None
        self.tmj_label.setText("TMJ PDF: Not selected")
        self.portfolio_label.setText("Portfolio Excel: Not selected")
        self.progress_label.setText("Ready")
        self.progress_bar.setValue(0)
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self.process_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.filter_input.clear()

    def update_progress(self, value, message):
        self.progress_bar.setValue(value)
        self.progress_label.setText(message)

    def show_error(self, message):
        QMessageBox.critical(self, "Error", message)
        self.progress_label.setText("Error Occurred")
        self.set_ui_processing_state(False)

    def display_results(self, results):
        self.results = results
        if results["Conflict Marks"].str.strip().eq("").all():
            QMessageBox.information(self, "Info", "No similar trademarks found.")
            self.progress_label.setText("No Conflicts Found")
            self.set_ui_processing_state(False)
            return

        self.table.setRowCount(len(results))
        self.table.setColumnCount(len(results.columns))
        self.table.setHorizontalHeaderLabels(results.columns)
        self.table.setFont(Styles.FONT)

        for row_idx, row in results.iterrows():
            for col_idx, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                if col_idx == results.columns.get_loc("Conflict Marks") and value:
                    item.setBackground(QColor(Styles.DARK_THEME["highlight"]))
                self.table.setItem(row_idx, col_idx, item)

        self.table.resizeColumnsToContents()
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.export_btn.setEnabled(True)
        self.set_ui_processing_state(False)

    def filter_table(self):
        filter_text = self.filter_input.text().lower()
        for row in range(self.table.rowCount()):
            match = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and filter_text in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(row, not match)

    def export_to_excel(self):
        if self.results is not None:
            try:
                wb = Workbook()
                ws_details = wb.active
                ws_details.title = "Conflict Details"

                # Title
                title = f"Trademark Similarity Conflict Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws_details['A1'] = title
                ws_details.merge_cells('A1:I1')
                ws_details['A1'].font = Font(bold=True, size=14)
                ws_details['A1'].alignment = Alignment(horizontal='center')
                ws_details['A1'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                # Headers
                headers = self.results.columns.tolist()
                for col, header in enumerate(headers, 1):
                    cell = ws_details.cell(row=2, column=col, value=header)
                    cell.font = Styles.HEADER_FONT
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')

                # Data
                for row in dataframe_to_rows(self.results, index=False, header=False):
                    ws_details.append(row)

                # Styling
                for row in ws_details[2:ws_details.max_row + 1]:
                    for cell in row:
                        cell.border = Styles.BORDER
                        if cell.column in [headers.index("File number") + 1, headers.index("Application number") + 1]:
                            cell.alignment = Alignment(horizontal='center')
                        if cell.column == headers.index("Other Details") + 1:
                            cell.alignment = Alignment(wrap_text=True)
                        if cell.column == headers.index("Conflict Marks") + 1 and cell.value:
                            cell.fill = PatternFill(start_color=Styles.DARK_THEME["highlight"], end_color=Styles.DARK_THEME["highlight"], fill_type="solid")

                # Status Formatting
                status_col = headers.index("Status") + 1
                for row in ws_details.iter_rows(min_row=3, max_row=ws_details.max_row, min_col=status_col, max_col=status_col):
                    for cell in row:
                        if "Registered" in cell.value:
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif "Closed" in cell.value:
                            cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
                        elif "Pending" in cell.value:
                            cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

                # Adjust Column Widths
                for col in range(1, len(headers) + 1):
                    ws_details.column_dimensions[ws_details.cell(row=2, column=col).column_letter].width = max(
                        len(str(ws_details.cell(row=r, column=col).value)) + 2 for r in range(2, ws_details.max_row + 1)
                    ) if ws_details.max_row > 2 else 15

                # Freeze Panes and Summary
                ws_details.freeze_panes = "A3"
                ws_details[f"A{ws_details.max_row + 2}"] = f"Total Portfolio Trademarks: {len(self.results)}"
                ws_details[f"B{ws_details.max_row}"] = f"Total Conflicts Identified: {sum(1 for x in self.results['Conflict Marks'] if x)}"
                ws_details[f"A{ws_details.max_row + 4}"] = "Confidential - For Internal Use Only"
                ws_details[f"A{ws_details.max_row + 4}"].alignment = Alignment(horizontal='center')
                ws_details.merge_cells(f"A{ws_details.max_row + 4}:I{ws_details.max_row + 4}")

                # Conflict Summary Sheet
                ws_summary = wb.create_sheet(title="Conflict Summary")
                ws_summary['A1'] = f"TMJ Conflict Summary - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws_summary.merge_cells('A1:B1')
                ws_summary['A1'].font = Font(bold=True, size=14)
                ws_summary['A1'].alignment = Alignment(horizontal='center')
                ws_summary['A1'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                conflict_data = []
                for idx, row in self.results.iterrows():
                    if row["Conflict Marks"]:
                        for mark in row["Conflict Marks"].split(", "):
                            conflict_data.append({"TMJ Trademark": mark.strip()})

                if not conflict_data:
                    ws_summary['A3'] = "No conflicts found."
                else:
                    conflict_df = pd.DataFrame(conflict_data)
                    pivot_data = conflict_df.groupby("TMJ Trademark").size().reset_index(name="Number of Conflicts")
                    for col, header in enumerate(["TMJ Trademark", "Number of Conflicts"], 1):
                        cell = ws_summary.cell(row=2, column=col, value=header)
                        cell.font = Styles.HEADER_FONT
                        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center')
                    for row_idx, row in enumerate(pivot_data.itertuples(), 3):
                        ws_summary.cell(row=row_idx, column=1, value=row._1)
                        ws_summary.cell(row=row_idx, column=2, value=row._2)

                    for row in ws_summary[2:ws_summary.max_row + 1]:
                        for cell in row:
                            cell.border = Styles.BORDER
                            cell.alignment = Alignment(horizontal='center')

                    for col in range(1, 3):
                        ws_summary.column_dimensions[ws_summary.cell(row=2, column=col).column_letter].width = max(
                            len(str(ws_summary.cell(row=r, column=col).value)) + 2 for r in range(2, ws_summary.max_row + 1)
                        ) if ws_summary.max_row > 2 else 15

                ws_summary[f"A{ws_summary.max_row + 2}"] = "Confidential - For Internal Use Only"
                ws_summary[f"A{ws_summary.max_row + 2}"].alignment = Alignment(horizontal='center')
                ws_summary.merge_cells(f"A{ws_summary.max_row + 2}:B{ws_summary.max_row + 2}")

                filename = f"Trademark_Conflict_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
                wb.save(filename)
                QMessageBox.information(self, "Success", f"Report exported as {filename}")
                logger.info(f"Report exported successfully as {filename}")

            except Exception as e:
                logger.error(f"Failed to export report: {e}")
                QMessageBox.critical(self, "Error", f"Failed to export report: {str(e)}")

# Main Execution
if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        window = TrademarkApp()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        logger.critical(f"Application failed to start: {e}")
        sys.exit(1)